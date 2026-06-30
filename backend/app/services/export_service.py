"""
导出服务
支持多种格式导出：
- Excel (.xlsx): 多sheet实验数据导出
- CSV: 已有
- JSON: 已有
- PNG/PDF/SVG图表导出
- 混淆矩阵导出
"""
import os
import io
import json
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

import numpy as np

logger = logging.getLogger(__name__)


def export_experiments_to_excel(
    experiments: List[Dict[str, Any]],
    metrics_map: Optional[Dict[str, List[Dict]]] = None,
) -> bytes:
    """导出实验列表为Excel格式，多sheet

    Args:
        experiments: 实验列表
        metrics_map: {experiment_id: [metrics...]} 训练指标

    Returns:
        Excel文件字节内容
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Sheet 1: 实验汇总
    ws_summary = wb.active
    ws_summary.title = "实验汇总"

    headers = [
        "实验ID", "名称", "描述", "状态", "模型", "参数量", "层数",
        "最佳准确率", "最终Loss", "Epoch数", "F1分数", "精确率", "召回率",
        "学习率", "批次大小", "优化器", "标签", "创建时间"
    ]
    ws_summary.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for exp in experiments:
        config = exp.get("config", {}) or {}
        hyperparams = exp.get("hyperparams", {}) or {}
        model_arch = exp.get("model_architecture", {}) or {}
        row = [
            exp.get("experiment_id", ""),
            exp.get("name", ""),
            exp.get("description", ""),
            exp.get("status", ""),
            config.get("model_architecture", model_arch.get("class_name", "")),
            exp.get("total_params", model_arch.get("total_params", 0)),
            exp.get("layer_count", model_arch.get("num_layers", 0)),
            round(config.get("best_val_accuracy", exp.get("best_accuracy", 0) or 0), 4),
            round(config.get("final_val_loss", exp.get("final_loss", 0) or 0), 6),
            exp.get("total_epochs", 0),
            round(config.get("f1", 0), 4),
            round(config.get("precision", 0), 4),
            round(config.get("recall", 0), 4),
            hyperparams.get("learning_rate", ""),
            hyperparams.get("batch_size", ""),
            hyperparams.get("optimizer", ""),
            ", ".join(exp.get("tags", []) or []),
            exp.get("created_at", ""),
        ]
        ws_summary.append(row)

    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_summary.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # Sheet 2: 训练指标详情（每个实验一个sheet或合并）
    if metrics_map:
        ws_metrics = wb.create_sheet("训练指标")
        metric_headers = ["实验ID", "Epoch", "Train Loss", "Val Loss", "Train Acc", "Val Acc",
                          "Precision", "Recall", "F1", "LR", "Grad Norm", "Weight Norm"]
        ws_metrics.append(metric_headers)
        for col_idx in range(1, len(metric_headers) + 1):
            cell = ws_metrics.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for exp_id, metrics in metrics_map.items():
            for m in metrics:
                extra = {}
                if m.get("extra_data"):
                    try:
                        extra = json.loads(m["extra_data"]) if isinstance(m["extra_data"], str) else m["extra_data"]
                    except:
                        pass
                ws_metrics.append([
                    exp_id,
                    m.get("epoch", m.get("step", 0)),
                    round(m.get("loss", 0), 6),
                    round(m.get("val_loss", 0), 6),
                    round(m.get("accuracy", 0), 4),
                    round(m.get("val_accuracy", 0), 4),
                    round(extra.get("precision", 0), 4),
                    round(extra.get("recall", 0), 4),
                    round(extra.get("f1", 0), 4),
                    m.get("learning_rate", 0),
                    round(extra.get("gradient_norm", 0), 4),
                    round(extra.get("weight_norm", 0), 4),
                ])

        for col in ws_metrics.columns:
            ws_metrics.column_dimensions[col[0].column_letter].width = 16

    # Sheet 3: 混淆矩阵（最后一个实验的混淆矩阵）
    if metrics_map:
        last_exp_id = list(metrics_map.keys())[-1]
        last_metrics = metrics_map[last_exp_id]
        if last_metrics:
            last_m = last_metrics[-1]
            extra = {}
            if last_m.get("extra_data"):
                try:
                    extra = json.loads(last_m["extra_data"]) if isinstance(last_m["extra_data"], str) else last_m["extra_data"]
                except:
                    pass
            cm = extra.get("confusion_matrix")
            if cm and isinstance(cm, list):
                ws_cm = wb.create_sheet("混淆矩阵")
                num_classes = len(cm)
                header_row = ["预测\\真实"] + [f"类别{i}" for i in range(num_classes)]
                ws_cm.append(header_row)
                for col_idx in range(1, len(header_row) + 1):
                    cell = ws_cm.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = thin_border

                for i, row in enumerate(cm):
                    ws_cm.append([f"类别{i}"] + row)

                for col in ws_cm.columns:
                    ws_cm.column_dimensions[col[0].column_letter].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def export_single_experiment_excel(
    experiment: Dict[str, Any],
    metrics: List[Dict],
) -> bytes:
    """导出单个实验的完整Excel报告"""
    return export_experiments_to_excel([experiment], {experiment.get("experiment_id", ""): metrics})


def export_confusion_matrix_csv(cm: List[List[int]], class_names: Optional[List[str]] = None) -> str:
    """导出混淆矩阵为CSV格式"""
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    n = len(cm)
    header = [""] + [class_names[i] if class_names and i < len(class_names) else f"Class_{i}" for i in range(n)]
    writer.writerow(header)
    for i, row in enumerate(cm):
        label = class_names[i] if class_names and i < len(class_names) else f"Class_{i}"
        writer.writerow([label] + row)
    output.seek(0)
    return output.getvalue()


def export_roc_data(fpr: List[float], tpr: List[float], auc_score: float, thresholds: Optional[List[float]] = None) -> str:
    """导出ROC曲线数据为CSV"""
    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([f"# AUC = {auc_score:.4f}"])
    writer.writerow(["FPR", "TPR", "Threshold"])
    for i in range(len(fpr)):
        t = thresholds[i] if thresholds and i < len(thresholds) else ""
        writer.writerow([fpr[i], tpr[i], t])
    output.seek(0)
    return output.getvalue()


def _get_cm(data: Dict[str, Any]) -> List[List[int]]:
    """从data中获取混淆矩阵，兼容多种key"""
    cm = data.get("cm") or data.get("matrix") or data.get("confusion_matrix") or []
    if isinstance(cm, np.ndarray):
        cm = cm.tolist()
    return cm

def _get_roc_data(data: Dict[str, Any]) -> Dict[str, Any]:
    """从data中获取ROC数据，兼容多种格式"""
    roc = data.get("roc_data") or data.get("roc_curve") or {}
    if not roc and "fpr" in data and "tpr" in data:
        roc = data
    return roc if isinstance(roc, dict) else {}

def _get_pr_data(data: Dict[str, Any]) -> Dict[str, Any]:
    """从data中获取PR数据，兼容多种格式"""
    pr = data.get("pr_data") or data.get("pr_curve") or {}
    if not pr and "precision" in data and "recall" in data:
        pr = data
    return pr if isinstance(pr, dict) else {}


def export_chart_png(
    chart_type: str,
    data: Dict[str, Any],
    dpi: int = 300,
    title: Optional[str] = None,
) -> bytes:
    """导出图表为高分辨率PNG（300DPI论文级）

    Args:
        chart_type: 图表类型 (loss_curve, acc_curve, confusion_matrix, roc, pr)
        data: 图表数据
        dpi: 分辨率，默认300
        title: 图表标题

    Returns:
        PNG文件字节
    """
    from app.ml.metrics import (
        plot_training_curves_base64,
        plot_confusion_matrix_base64,
        plot_roc_curve_base64,
        plot_pr_curve_base64,
    )
    import base64

    if chart_type in ("loss_curve", "acc_curve"):
        result = plot_training_curves_base64(data.get("metrics", []), dpi=dpi)
        b64_str = result.get(chart_type, "")
    elif chart_type == "confusion_matrix":
        cm = _get_cm(data)
        b64_str = plot_confusion_matrix_base64(
            cm,
            class_names=data.get("class_names"),
            dpi=dpi,
            normalize=data.get("normalize", False),
        )
    elif chart_type == "roc":
        roc = _get_roc_data(data)
        b64_str = plot_roc_curve_base64(roc, dpi=dpi)
    elif chart_type == "pr":
        pr = _get_pr_data(data)
        b64_str = plot_pr_curve_base64(pr, dpi=dpi)
    else:
        raise ValueError(f"不支持的图表类型: {chart_type}")

    return base64.b64decode(b64_str)


def export_chart_svg(
    chart_type: str,
    data: Dict[str, Any],
    title: Optional[str] = None,
) -> bytes:
    """导出图表为SVG矢量格式

    Args:
        chart_type: 图表类型
        data: 图表数据
        title: 标题

    Returns:
        SVG文件字节
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt

    fig = None

    if chart_type in ("loss_curve", "acc_curve"):
        metrics = data.get("metrics", [])
        epochs = [m.get("epoch", m.get("step", 0)) for m in metrics]
        fig, ax = plt.subplots(1, 1, figsize=(8, 5))
        if chart_type == "loss_curve":
            ax.plot(epochs, [m.get("loss", 0) for m in metrics], 'b-', lw=2, label='Train Loss')
            ax.plot(epochs, [m.get("val_loss", 0) for m in metrics], 'r-', lw=2, label='Val Loss')
            ax.set_ylabel('Loss')
            ax.set_title(title or 'Training & Validation Loss')
        else:
            ax.plot(epochs, [m.get("accuracy", 0) for m in metrics], 'b-', lw=2, label='Train Acc')
            ax.plot(epochs, [m.get("val_accuracy", 0) for m in metrics], 'r-', lw=2, label='Val Acc')
            ax.set_ylabel('Accuracy')
            ax.set_ylim([0, 1.05])
            ax.set_title(title or 'Training & Validation Accuracy')
        ax.set_xlabel('Epoch')
        ax.legend(fontsize=10)
        ax.grid(True, alpha=0.3)

    elif chart_type == "confusion_matrix":
        cm_raw = _get_cm(data)
        cm = np.array(cm_raw, dtype=np.float64)
        class_names = data.get("class_names")
        n = cm.shape[0] if cm.ndim == 2 else 0
        if n == 0:
            raise ValueError("混淆矩阵数据为空或格式无效")
        if class_names is None:
            class_names = [f"Class {i}" for i in range(n)]
        if data.get("normalize", False):
            row_sums = cm.sum(axis=1, keepdims=True)
            cm = np.divide(cm, row_sums, where=row_sums != 0, out=np.zeros_like(cm))
        fig, ax = plt.subplots(1, 1, figsize=(max(6, n * 0.8), max(5, n * 0.7)))
        im = ax.imshow(cm, interpolation='nearest', cmap=plt.cm.Blues)
        fig.colorbar(im, ax=ax)
        ax.set(xticks=np.arange(n), yticks=np.arange(n),
               xticklabels=class_names, yticklabels=class_names,
               ylabel='True label', xlabel='Predicted label',
               title=title or 'Confusion Matrix')
        plt.setp(ax.get_xticklabels(), rotation=45, ha="right", rotation_mode="anchor")
        fmt = '.2f' if data.get("normalize", False) else '.0f'
        thresh = cm.max() / 2. if cm.max() > 0 else 0.5
        for i in range(n):
            for j in range(n):
                ax.text(j, i, format(cm[i, j], fmt), ha="center", va="center",
                        color="white" if cm[i, j] > thresh else "black", fontsize=9)
        fig.tight_layout()

    elif chart_type == "roc":
        roc_data = _get_roc_data(data)
        num_classes = roc_data.get("num_classes", 0)
        if num_classes == 0 and "fpr" in roc_data:
            num_classes = len(roc_data["fpr"])
        colors = plt.cm.tab10(np.linspace(0, 1, max(num_classes, 1)))
        fig, ax = plt.subplots(1, 1, figsize=(8, 6))
        for idx, (cls_name, fpr) in enumerate(roc_data.get("fpr", {}).items()):
            tpr = roc_data.get("tpr", {}).get(cls_name, [])
            auc_val = roc_data.get("auc_scores", {}).get(cls_name, 0)
            ax.plot(fpr, tpr, color=colors[idx % 10], lw=2, label=f'{cls_name} (AUC={auc_val:.4f})')
        ax.plot([0, 1], [0, 1], 'k--', lw=1, alpha=0.5)
        ax.set_xlim([0, 1]); ax.set_ylim([0, 1.05])
        ax.set_xlabel('False Positive Rate'); ax.set_ylabel('True Positive Rate')
        ax.set_title(title or f'ROC Curve (Macro AUC={roc_data.get("macro_auc", 0):.4f})')
        ax.legend(loc="lower right", fontsize=9); ax.grid(True, alpha=0.3)

    elif chart_type == "pr":
        pr_data = _get_pr_data(data)
        num_classes = pr_data.get("num_classes", 0)
        if num_classes == 0 and "precision" in pr_data:
            num_classes = len(pr_data["precision"])
        colors = plt.cm.tab10(np.linspace(0, 1, max(num_classes, 1)))
        fig, ax = plt.subplots(1, 1, figsize=(8, 6))
        for idx, (cls_name, recall) in enumerate(pr_data.get("recall", {}).items()):
            precision = pr_data.get("precision", {}).get(cls_name, [])
            ap_val = pr_data.get("ap_scores", {}).get(cls_name, 0)
            ax.plot(recall, precision, color=colors[idx % 10], lw=2, label=f'{cls_name} (AP={ap_val:.4f})')
        ax.set_xlim([0, 1]); ax.set_ylim([0, 1.05])
        ax.set_xlabel('Recall'); ax.set_ylabel('Precision')
        ax.set_title(title or f'PR Curve (Macro AP={pr_data.get("macro_ap", 0):.4f})')
        ax.legend(loc="lower left", fontsize=9); ax.grid(True, alpha=0.3)

    if fig is None:
        raise ValueError(f"无法生成图表类型: {chart_type}")

    buf = io.BytesIO()
    fig.savefig(buf, format='svg', bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def export_chart_pdf(
    chart_type: str,
    data: Dict[str, Any],
    dpi: int = 300,
    title: Optional[str] = None,
) -> bytes:
    """导出图表为PDF矢量格式（论文级）

    Args:
        chart_type: 图表类型
        data: 图表数据
        dpi: 分辨率（仅对嵌入的文本/线条清晰度有影响）
        title: 标题

    Returns:
        PDF文件字节（矢量格式）
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    buf = io.BytesIO()
    fig = None

    if chart_type in ("loss_curve", "acc_curve"):
        metrics = data.get("metrics", [])
        epochs = [m.get("epoch", m.get("step", 0)) for m in metrics]
        fig, ax = plt.subplots(1, 1, figsize=(8, 5))
        if chart_type == "loss_curve":
            ax.plot(epochs, [m.get("loss", 0) for m in metrics], 'b-', lw=2, label='Train Loss')
            ax.plot(epochs, [m.get("val_loss", 0) for m in metrics], 'r-', lw=2, label='Val Loss')
            ax.set_ylabel('Loss')
            ax.set_title(title or 'Training & Validation Loss')
        else:
            ax.plot(epochs, [m.get("accuracy", 0) for m in metrics], 'b-', lw=2, label='Train Acc')
            ax.plot(epochs, [m.get("val_accuracy", 0) for m in metrics], 'r-', lw=2, label='Val Acc')
            ax.set_ylabel('Accuracy')
            ax.set_ylim([0, 1.05])
            ax.set_title(title or 'Training & Validation Accuracy')
        ax.set_xlabel('Epoch')
        ax.legend(fontsize=10)
        ax.grid(True, alpha=0.3)

    elif chart_type == "confusion_matrix":
        cm_raw = _get_cm(data)
        cm = np.array(cm_raw, dtype=np.float64)
        class_names = data.get("class_names")
        n = cm.shape[0] if cm.ndim == 2 else 0
        if n == 0:
            raise ValueError("混淆矩阵数据为空或格式无效")
        if class_names is None:
            class_names = [f"Class {i}" for i in range(n)]
        if data.get("normalize", False):
            row_sums = cm.sum(axis=1, keepdims=True)
            cm = np.divide(cm, row_sums, where=row_sums != 0, out=np.zeros_like(cm))
        fig, ax = plt.subplots(1, 1, figsize=(max(6, n * 0.8), max(5, n * 0.7)))
        im = ax.imshow(cm, interpolation='nearest', cmap=plt.cm.Blues)
        fig.colorbar(im, ax=ax)
        ax.set(xticks=np.arange(n), yticks=np.arange(n),
               xticklabels=class_names, yticklabels=class_names,
               ylabel='True label', xlabel='Predicted label',
               title=title or 'Confusion Matrix')
        plt.setp(ax.get_xticklabels(), rotation=45, ha="right", rotation_mode="anchor")
        fmt = '.2f' if data.get("normalize", False) else '.0f'
        thresh = cm.max() / 2. if cm.max() > 0 else 0.5
        for i in range(n):
            for j in range(n):
                ax.text(j, i, format(cm[i, j], fmt), ha="center", va="center",
                        color="white" if cm[i, j] > thresh else "black", fontsize=9)
        fig.tight_layout()

    elif chart_type == "roc":
        roc_data = _get_roc_data(data)
        num_classes = roc_data.get("num_classes", 0)
        if num_classes == 0 and "fpr" in roc_data:
            num_classes = len(roc_data["fpr"])
        colors = plt.cm.tab10(np.linspace(0, 1, max(num_classes, 1)))
        fig, ax = plt.subplots(1, 1, figsize=(8, 6))
        for idx, (cls_name, fpr) in enumerate(roc_data.get("fpr", {}).items()):
            tpr = roc_data.get("tpr", {}).get(cls_name, [])
            auc_val = roc_data.get("auc_scores", {}).get(cls_name, 0)
            ax.plot(fpr, tpr, color=colors[idx % 10], lw=2, label=f'{cls_name} (AUC={auc_val:.4f})')
        ax.plot([0, 1], [0, 1], 'k--', lw=1, alpha=0.5)
        ax.set_xlim([0, 1]); ax.set_ylim([0, 1.05])
        ax.set_xlabel('False Positive Rate'); ax.set_ylabel('True Positive Rate')
        ax.set_title(title or f'ROC Curve (Macro AUC={roc_data.get("macro_auc", 0):.4f})')
        ax.legend(loc="lower right", fontsize=9); ax.grid(True, alpha=0.3)

    elif chart_type == "pr":
        pr_data = _get_pr_data(data)
        num_classes = pr_data.get("num_classes", 0)
        if num_classes == 0 and "precision" in pr_data:
            num_classes = len(pr_data["precision"])
        colors = plt.cm.tab10(np.linspace(0, 1, max(num_classes, 1)))
        fig, ax = plt.subplots(1, 1, figsize=(8, 6))
        for idx, (cls_name, recall) in enumerate(pr_data.get("recall", {}).items()):
            precision = pr_data.get("precision", {}).get(cls_name, [])
            ap_val = pr_data.get("ap_scores", {}).get(cls_name, 0)
            ax.plot(recall, precision, color=colors[idx % 10], lw=2, label=f'{cls_name} (AP={ap_val:.4f})')
        ax.set_xlim([0, 1]); ax.set_ylim([0, 1.05])
        ax.set_xlabel('Recall'); ax.set_ylabel('Precision')
        ax.set_title(title or f'PR Curve (Macro AP={pr_data.get("macro_ap", 0):.4f})')
        ax.legend(loc="lower left", fontsize=9); ax.grid(True, alpha=0.3)

    if fig is None:
        raise ValueError(f"无法生成图表类型: {chart_type}")

    with PdfPages(buf) as pdf:
        pdf.savefig(fig, bbox_inches='tight', facecolor='white')
    plt.close(fig)

    buf.seek(0)
    return buf.getvalue()


def export_confusion_matrix_json(cm: List[List[int]], class_names: Optional[List[str]] = None) -> str:
    """导出混淆矩阵为JSON格式"""
    n = len(cm)
    result = {
        "matrix": cm,
        "class_names": class_names or [f"Class_{i}" for i in range(n)],
        "num_classes": n,
        "exported_at": datetime.utcnow().isoformat(),
    }
    return json.dumps(result, ensure_ascii=False, indent=2)


def export_metrics_json(metrics: List[Dict], roc_data: Optional[Dict] = None, pr_data: Optional[Dict] = None) -> str:
    """导出完整训练指标为JSON格式，包含ROC/PR数据"""
    result = {
        "metrics": metrics,
        "roc_curve": roc_data,
        "pr_curve": pr_data,
        "exported_at": datetime.utcnow().isoformat(),
    }
    return json.dumps(result, ensure_ascii=False, indent=2)


def export_pytorch_to_onnx(experiment_dir: str, input_shape: tuple = (1, 1, 28, 28)) -> bytes:
    """将训练好的PyTorch模型导出为ONNX格式

    Args:
        experiment_dir: 实验模型目录（包含best.pt）
        input_shape: 模型输入形状 (batch, channels, height, width)，仅作fallback

    Returns:
        ONNX模型文件字节内容
    """
    import torch
    from app.ml.model_builder import build_model

    model_dir = Path(experiment_dir) if not isinstance(experiment_dir, Path) else experiment_dir
    best_path = model_dir / "best.pt"
    last_path = model_dir / "last.pt"

    checkpoint_path = best_path if best_path.exists() else (last_path if last_path.exists() else None)
    if checkpoint_path is None:
        raise FileNotFoundError(f"在 {model_dir} 中未找到模型检查点(best.pt/last.pt)")

    raw_checkpoint = torch.load(str(checkpoint_path), map_location="cpu", weights_only=False)

    state_dict = raw_checkpoint
    if isinstance(raw_checkpoint, dict) and ("model_state_dict" in raw_checkpoint or "state_dict" in raw_checkpoint):
        state_dict = raw_checkpoint.get("model_state_dict", raw_checkpoint.get("state_dict", raw_checkpoint))

    from app.core.database import SessionLocal
    from app.models.experiment import Experiment as ExperimentModel

    dataset_type = "numpy"
    feature_shape_str = "1x28x28"
    num_classes = 10
    exp_model_config = {}

    db = SessionLocal()
    try:
        exp_id = model_dir.name
        exp = db.query(ExperimentModel).filter(ExperimentModel.experiment_id == exp_id).first()
        if exp:
            if exp.model_architecture:
                try:
                    arch = json.loads(exp.model_architecture) if isinstance(exp.model_architecture, str) else exp.model_architecture
                    arch_type = arch.get("type", "")
                    if arch_type in ("image_folder", "mnist_idx", "csv", "numpy"):
                        dataset_type = arch_type
                    ishape = arch.get("input_shape", None)
                    if isinstance(ishape, (list, tuple)) and len(ishape) >= 3:
                        c, h, w = int(ishape[0]), int(ishape[1]), int(ishape[2])
                        if c <= 4 and h > 4 and w > 4:
                            feature_shape_str = f"{c}x{h}x{w}"
                        elif w <= 4 and h > 4 and c > 4:
                            feature_shape_str = f"{w}x{h}x{c}"
                    if arch.get("num_classes"):
                        num_classes = int(arch["num_classes"])
                except Exception:
                    pass

            if exp.config:
                try:
                    cfg = json.loads(exp.config) if isinstance(exp.config, str) else exp.config
                    inner_cfg = cfg.get("model_config", cfg.get("config", cfg))
                    if isinstance(inner_cfg, dict):
                        exp_model_config.update(inner_cfg)
                    for k, v in cfg.items():
                        if k in ("channels", "attention", "use_bn", "use_dropout", "dropout_rate", "use_residual", "fc_hidden", "use_attention"):
                            exp_model_config[k] = v
                except Exception:
                    pass

            if exp.model_id:
                try:
                    from app.services.dataset_service import DatasetService
                    ds_svc = DatasetService()
                    ds_info = ds_svc.get_dataset(exp.model_id)
                    if ds_info:
                        ds_data = ds_svc.load_dataset(exp.model_id, split="test")
                        dataset_type = ds_data.get("dataset_type", dataset_type)
                        ds_fs = ds_data.get("feature_shape", "")
                        if ds_fs and isinstance(ds_fs, str) and "x" in ds_fs:
                            feature_shape_str = ds_fs
                        num_classes = len(ds_data.get("label_names", [])) or num_classes
                except Exception:
                    pass
    finally:
        db.close()

    model_cfg = {
        "channels": exp_model_config.get("channels", exp_model_config.get("channel_list", [32, 64])),
        "use_bn": exp_model_config.get("use_bn", True),
        "use_dropout": exp_model_config.get("use_dropout", True),
        "dropout_rate": exp_model_config.get("dropout_rate", 0.3),
        "attention": exp_model_config.get("attention", "none"),
        "use_residual": exp_model_config.get("use_residual", False),
        "fc_hidden": exp_model_config.get("fc_hidden", 128),
        "use_attention": exp_model_config.get("use_attention", False),
    }

    try:
        model = build_model(
            dataset_type=dataset_type,
            feature_shape=feature_shape_str,
            num_classes=num_classes,
            model_config=model_cfg,
        )
    except Exception as e:
        logger.warning(f"使用config构建模型失败({e}), 尝试默认配置")
        model = build_model(
            dataset_type=dataset_type,
            feature_shape=feature_shape_str,
            num_classes=num_classes,
        )

    try:
        model.load_state_dict(state_dict, strict=False)
    except Exception as e:
        logger.warning(f"加载state_dict警告(可能不影响推理): {e}")

    model.eval()

    parts = feature_shape_str.split("x")
    c, h, w = int(parts[0]), int(parts[1]), int(parts[2])
    actual_input_shape = (input_shape[0], c, h, w)
    dummy_input = torch.randn(*actual_input_shape)

    buf = io.BytesIO()
    try:
        with torch.no_grad():
            torch.onnx.export(
                model,
                dummy_input,
                buf,
                export_params=True,
                opset_version=17,
                do_constant_folding=True,
                input_names=["input"],
                output_names=["output"],
                dynamic_axes={
                    "input": {0: "batch_size"},
                    "output": {0: "batch_size"},
                },
            )
    except Exception as e:
        raise RuntimeError(f"ONNX导出失败: {str(e)}")

    buf.seek(0)
    return buf.getvalue()
